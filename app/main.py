import os
import io
import re
import json
import hashlib
import sqlite3
import subprocess
import tempfile
from datetime import datetime, timedelta, time
from dateutil import tz

from openpyxl import load_workbook

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/calendar.events",
]

DUBLIN_TZ = tz.gettz("Europe/Dublin")


def env(name: str, default: str | None = None) -> str:
    val = os.getenv(name, default)
    if val is None or val == "":
        raise RuntimeError(f"Missing required env var: {name}")
    return val


def init_db(db_path: str) -> None:
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS processed_messages (
                message_id TEXT PRIMARY KEY,
                attachment_sha256 TEXT NOT NULL,
                processed_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS shift_events (
                shift_key TEXT PRIMARY KEY,
                event_id TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
def ensure_xlsx_on_disk(saved_path: str) -> str:
    if saved_path.lower().endswith(".xlsx"):
        return saved_path

    if saved_path.lower().endswith(".xls"):
        # Convert .xls -> .xlsx using libreoffice
        out_dir = os.path.dirname(saved_path)
        subprocess.run(
            [
                "/Applications/LibreOffice.app/Contents/MacOS/soffice",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                out_dir,
                saved_path,
            ],
            check=True,
        )
        converted = os.path.splitext(saved_path)[0] + ".xlsx"
        return converted

    raise RuntimeError(f"Unsupported attachment type: {saved_path}")

def db_has_message(db_path: str, message_id: str, sha: str) -> bool:
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            "SELECT 1 FROM processed_messages WHERE message_id=? AND attachment_sha256=?",
            (message_id, sha),
        ).fetchone()
        return row is not None


def db_mark_message(db_path: str, message_id: str, sha: str) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "INSERT OR REPLACE INTO processed_messages(message_id, attachment_sha256, processed_at) VALUES (?,?,?)",
            (message_id, sha, datetime.utcnow().isoformat()),
        )


def db_get_event_id(db_path: str, shift_key: str) -> str | None:
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            "SELECT event_id FROM shift_events WHERE shift_key=?",
            (shift_key,),
        ).fetchone()
        return row[0] if row else None


def db_upsert_event_id(db_path: str, shift_key: str, event_id: str) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "INSERT OR REPLACE INTO shift_events(shift_key, event_id, updated_at) VALUES (?,?,?)",
            (shift_key, event_id, datetime.utcnow().isoformat()),
        )


def load_credentials(creds_json_path: str, token_path: str) -> Credentials:
    # Token is stored on a mounted volume so CronJob runs reuse it.
    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            # google-auth refresh happens automatically in requests via googleapiclient
            pass
        else:
            # First-time interactive auth (run locally once outside k8s to generate token)
            flow = InstalledAppFlow.from_client_secrets_file(creds_json_path, SCOPES)
            creds = flow.run_local_server(port=0)

        os.makedirs(os.path.dirname(token_path), exist_ok=True)
        with open(token_path, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    return creds


def gmail_find_latest_with_attachment(gmail, query: str) -> dict | None:
    resp = gmail.users().messages().list(userId="me", q=query, maxResults=5).execute()
    msgs = resp.get("messages", [])
    if not msgs:
        return None
    # First result is usually newest, but we’ll still fetch metadata for safety if needed.
    return msgs[0]


def gmail_download_first_excel_attachment(gmail, message_id: str) -> tuple[bytes, str]:
    msg = gmail.users().messages().get(userId="me", id=message_id).execute()
    parts = msg.get("payload", {}).get("parts", [])
    if not parts:
        raise RuntimeError("Email has no parts/attachments.")

    # Walk parts; find first xls/xlsx attachment.
    def walk(p):
        yield p
        for sp in p.get("parts", []) or []:
            yield from walk(sp)

    filename = None
    att_id = None
    for p in walk({"parts": parts}):
        fn = p.get("filename")
        body = p.get("body", {})
        aid = body.get("attachmentId")
        if fn and aid and re.search(r"\.(xlsx|xls)$", fn, re.I):
            filename, att_id = fn, aid
            break

    if not att_id:
        raise RuntimeError("No .xls/.xlsx attachment found in the email.")

    att = (
        gmail.users()
        .messages()
        .attachments()
        .get(userId="me", messageId=message_id, id=att_id)
        .execute()
    )
    data = att.get("data")
    if not data:
        raise RuntimeError("Attachment download failed (no data).")

    import base64

    raw = base64.urlsafe_b64decode(data.encode("utf-8"))
    return raw, filename


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def parse_schedule_from_excel(excel_bytes: bytes, your_name: str) -> list[dict]:
    """
    Robust parser:
    - Finds header row by scanning first 30 rows
    - Detects name column by matching Name/Employee/etc.
    - Detects day columns by date pattern dd/mm/yy in header cells
    - Extracts start/end times from two columns per day
    """
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    def norm(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return str(v).strip()

    def looks_like_date_header(v: str) -> bool:
        return bool(re.search(r"\b\d{2}/\d{2}/\d{2}\b", v))

    def parse_date_from_header(v: str):
        m = re.search(r"(\d{2})/(\d{2})/(\d{2})", v)
        if not m:
            return None
        dd, mm, yy = map(int, m.groups())
        return datetime(2000 + yy, mm, dd).date()

    # 1) Find the header row
    header_row_idx = None
    header_values = None

    NAME_CANDIDATES = {"name", "employee", "colleague", "associate", "staff"}

    max_scan = min(ws.max_row, 30)
    for r in range(1, max_scan + 1):
        row = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        row_lower = [x.lower() for x in row]

        has_name_header = any(x in NAME_CANDIDATES for x in row_lower)
        has_any_date = any(looks_like_date_header(x) for x in row)

        # We consider it a header if it has a name-like cell and at least one date header
        if has_name_header and has_any_date:
            header_row_idx = r
            header_values = row
            break

    if header_row_idx is None:
        # Helpful debug
        sample = []
        for r in range(1, min(ws.max_row, 10) + 1):
            row = [norm(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 12) + 1)]
            sample.append((r, row))
        raise RuntimeError(
            "Could not find header row automatically. "
            "Top rows sample (row_num, first_12_cells): "
            f"{sample}"
        )

    # 2) Find the name column
    header_lower = [x.lower() for x in header_values]
    name_col = None
    for i, x in enumerate(header_lower):
        if x in NAME_CANDIDATES:
            name_col = i + 1
            break
    if not name_col:
        raise RuntimeError(f"Found header row {header_row_idx} but could not find a Name/Employee column.")

    # 3) Build day columns (each day has two columns: start/end)
    day_cols: list[tuple[datetime.date, int, int]] = []
    c = name_col + 1
    while c <= ws.max_column:
        hv = header_values[c - 1] if c - 1 < len(header_values) else ""
        if isinstance(hv, str) and looks_like_date_header(hv):
            day_date = parse_date_from_header(hv)
            if day_date:
                day_cols.append((day_date, c, c + 1))
                c += 2
                continue
        c += 1

    if not day_cols:
        raise RuntimeError(f"Header row {header_row_idx} found but no day columns detected.")

    # 4) Find your row (start scanning after header row)
    your_row = None
    for r in range(header_row_idx + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        if isinstance(name, str) and name.strip() == your_name:
            your_row = r
            break

    if your_row is None:
        # fallback: case-insensitive contains
        for r in range(header_row_idx + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            if isinstance(name, str) and your_name.lower() in name.lower():
                your_row = r
                break

    if your_row is None:
        raise RuntimeError(
            f"Could not find your name '{your_name}' under the Name column (header row {header_row_idx})."
        )

    # 5) Extract shifts
    def to_time(v) -> time:
        if isinstance(v, time):
            return v
        if isinstance(v, datetime):
            return v.time()
        if isinstance(v, str):
            v = v.strip()
            if not v:
                raise ValueError("Empty time string")
            hh, mm = map(int, v.split(":"))
            return time(hh, mm)
        # Sometimes excel stores time as float fraction of day
        if isinstance(v, (int, float)):
            # Excel time: fraction of a day
            total_minutes = int(round(float(v) * 24 * 60))
            hh = total_minutes // 60
            mm = total_minutes % 60
            return time(hh % 24, mm)
        raise RuntimeError(f"Unrecognized time value: {v!r}")

    shifts = []
    for day_date, sc, ec in day_cols:
        start_val = ws.cell(row=your_row, column=sc).value
        end_val = ws.cell(row=your_row, column=ec).value

        def is_blank(v) -> bool:
            if v is None:
                return True
            if isinstance(v, str) and v.strip() == "":
                return True
            return False

        if is_blank(start_val) or is_blank(end_val):
            continue

        st = to_time(start_val)
        et = to_time(end_val)        

        start_dt = datetime.combine(day_date, st).replace(tzinfo=DUBLIN_TZ)
        end_dt = datetime.combine(day_date, et).replace(tzinfo=DUBLIN_TZ)
        if end_dt <= start_dt:
            end_dt = end_dt + timedelta(days=1)

        shifts.append({"date": day_date, "start_dt": start_dt, "end_dt": end_dt})

    return shifts


def upsert_calendar_events(calendar, db_path: str, calendar_id: str, shifts: list[dict]) -> None:
    for s in shifts:
        start_dt: datetime = s["start_dt"]
        end_dt: datetime = s["end_dt"]

        # shift_key should be stable
        shift_key = f"{start_dt.isoformat()}|{end_dt.isoformat()}|{calendar_id}"
        existing_event_id = db_get_event_id(db_path, shift_key)

        body = {
            "summary": "Penneys/Primark Shift",
            "description": "Auto-imported from schedule email attachment.",
            "start": {"dateTime": start_dt.isoformat(), "timeZone": "Europe/Dublin"},
            "end": {"dateTime": end_dt.isoformat(), "timeZone": "Europe/Dublin"},
        }

        try:
            if existing_event_id:
                ev = calendar.events().update(
                    calendarId=calendar_id,
                    eventId=existing_event_id,
                    body=body,
                ).execute()
                db_upsert_event_id(db_path, shift_key, ev["id"])
            else:
                ev = calendar.events().insert(
                    calendarId=calendar_id,
                    body=body,
                ).execute()
                db_upsert_event_id(db_path, shift_key, ev["id"])
        except HttpError as e:
            raise RuntimeError(f"Calendar API error: {e}")


def main():
    your_name = env("YOUR_NAME")
    gmail_query = env(
        "GMAIL_QUERY",
        'has:attachment (filename:xlsx OR filename:xls) newer_than:60d',
    )
    calendar_id = env("CALENDAR_ID", "primary")

    data_dir = env("DATA_DIR", "/data")
    creds_json_path = env("GOOGLE_CREDS_JSON", "/secrets/credentials.json")
    token_path = os.path.join(data_dir, "token.json")
    db_path = os.path.join(data_dir, "state.db")

    init_db(db_path)

    creds = load_credentials(creds_json_path, token_path)
    gmail = build("gmail", "v1", credentials=creds)
    cal = build("calendar", "v3", credentials=creds)

    latest = gmail_find_latest_with_attachment(gmail, gmail_query)
    if not latest:
        print("No matching email found.")
        return

    message_id = latest["id"]
    excel_bytes, filename = gmail_download_first_excel_attachment(gmail, message_id)
    sha = sha256_bytes(excel_bytes)
    # DEBUG: save to disk so we can inspect it
    os.makedirs(os.path.join(data_dir, "attachments"), exist_ok=True)
    saved_path = os.path.join(data_dir, "attachments", filename or "attachment.bin")
    with open(saved_path, "wb") as f:
        f.write(excel_bytes)

    print("Saved attachment to:", saved_path)
    print("First 8 bytes:", excel_bytes[:8])
    xlsx_path = ensure_xlsx_on_disk(saved_path)

    with open(xlsx_path, "rb") as f:
        excel_bytes = f.read()

    shifts = parse_schedule_from_excel(excel_bytes, your_name)

    if db_has_message(db_path, message_id, sha):
        print("Already processed this email + attachment checksum. Exiting.")
        return

    shifts = parse_schedule_from_excel(excel_bytes, your_name)
    if not shifts:
        print("No shifts found for your name in the sheet.")
        db_mark_message(db_path, message_id, sha)
        return

    upsert_calendar_events(cal, db_path, calendar_id, shifts)
    db_mark_message(db_path, message_id, sha)

    print(f"Imported {len(shifts)} shifts from {filename} into calendar {calendar_id}.")


if __name__ == "__main__":
    main()